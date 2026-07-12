"""
chat Lambda
-------------------------------------------------------------
POST /chat

Handles:
  - action="list_files" → list objects for the requesting user
  - s3_keys=[...]       → combined multi-file question
  - s3_key=...          → single-file question

AI backend: Amazon Bedrock (Claude) — invoked via the Lambda's own IAM
role, no third-party API key. Patient data never leaves AWS's HIPAA
BAA boundary (unlike the previous OpenRouter-based implementation).
"""

import io
import json
import os

import boto3
from botocore.config import Config
from botocore.exceptions import ClientError

BUCKET_NAME      = os.environ.get("BUCKET_NAME",    "")
BUCKET_REGION    = os.environ.get("BUCKET_REGION",  "us-east-2")
ALLOWED_ORIGIN   = os.environ.get("ALLOWED_ORIGIN", "https://kbuddhiai.com")
BEDROCK_REGION   = os.environ.get("BEDROCK_REGION", "us-east-2")
BEDROCK_MODEL_ID = os.environ.get("BEDROCK_MODEL_ID", "us.anthropic.claude-sonnet-4-5-20250929-v1:0")

CORS_HEADERS = {
    "Access-Control-Allow-Origin":  ALLOWED_ORIGIN,
    "Access-Control-Allow-Headers": "Content-Type",
    "Access-Control-Allow-Methods": "POST, OPTIONS",
}


# ── File text extraction ───────────────────────────────────────────────────────

def extract_text(file_bytes: bytes, filename: str) -> str:
    ext = filename.lower().rsplit(".", 1)[-1]

    if ext in ("xlsx", "xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            lines = []
            for name in wb.sheetnames:
                ws = wb[name]
                lines.append(f"=== Sheet: {name} ===")
                for row in ws.iter_rows(values_only=True):
                    lines.append("\t".join("" if c is None else str(c) for c in row))
            return "\n".join(lines)
        except Exception:
            import xlrd
            wb = xlrd.open_workbook(file_contents=file_bytes)
            lines = []
            for sheet in wb.sheets():
                lines.append(f"=== Sheet: {sheet.name} ===")
                for i in range(sheet.nrows):
                    lines.append("\t".join(str(sheet.cell_value(i, j)) for j in range(sheet.ncols)))
            return "\n".join(lines)

    elif ext == "pdf":
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(file_bytes))
        lines  = []
        for i, page in enumerate(reader.pages, 1):
            t = page.extract_text()
            if t:
                lines.append(f"--- Page {i} ---\n{t}")
        return "\n".join(lines) or "[No extractable text found in PDF]"

    elif ext == "docx":
        from docx import Document
        doc = Document(io.BytesIO(file_bytes))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())

    elif ext in ("csv", "txt"):
        return file_bytes.decode("utf-8", errors="replace")

    else:
        try:
            return file_bytes.decode("utf-8", errors="replace")
        except Exception:
            return "[Could not extract text from this file type]"


# ── Bedrock (Claude) call ──────────────────────────────────────────────────────

MAX_FILE_CHARS = 450000

def call_bedrock(file_text: str, question: str, chat_history: list) -> str:
    truncated = file_text[:MAX_FILE_CHARS]
    if len(file_text) > MAX_FILE_CHARS:
        truncated += "\n\n[... file truncated for context length ...]"

    system_msg = (
        "You are a helpful assistant. The user has uploaded a file whose content is below. "
        "Answer questions about it accurately and concisely.\n\n"
        f"FILE CONTENT:\n{truncated}"
    )

    # Converse API: system prompt is a separate top-level param, not a
    # message with role "system". chat_history is already {role, content}
    # pairs alternating user/assistant, which is what Converse expects —
    # just needs each "content" wrapped as [{"text": ...}].
    messages = []
    for msg in chat_history:
        messages.append({"role": msg["role"], "content": [{"text": msg["content"]}]})
    messages.append({"role": "user", "content": [{"text": question}]})

    bedrock = boto3.client("bedrock-runtime", region_name=BEDROCK_REGION)
    response = bedrock.converse(
        modelId=BEDROCK_MODEL_ID,
        system=[{"text": system_msg}],
        messages=messages,
        inferenceConfig={"maxTokens": 32000},
    )
    return response["output"]["message"]["content"][0]["text"]


# ── Lambda handler ─────────────────────────────────────────────────────────────

def lambda_handler(event, context):
    if event.get("httpMethod") == "OPTIONS":
        return {"statusCode": 200, "headers": CORS_HEADERS, "body": ""}

    # Lambda Function URLs (payload format v2.0) already inject their own
    # Access-Control-Allow-Origin header per the CORS config on the URL
    # resource. Adding our own on top of that produces two conflicting
    # headers, which browsers reject. Only add ours for the API Gateway
    # route (v1 proxy integration), which has no such built-in handling.
    is_function_url = event.get("version") == "2.0"

    def resp(status, body):
        return _resp(status, body, {} if is_function_url else CORS_HEADERS)

    try:
        body         = json.loads(event.get("body") or "{}")
        action       = body.get("action", "")
        s3_key       = body.get("s3_key", "").strip()
        s3_keys      = body.get("s3_keys", [])
        question     = body.get("question", "").strip()
        chat_history = body.get("chat_history", [])
        user_sub     = body.get("user_sub", "").strip()  # optional filter for list_files

        s3_client = boto3.client(
            "s3",
            region_name=BUCKET_REGION,
            config=Config(s3={"addressing_style": "virtual"}),
        )

        # ── List files (filtered to the requesting user when user_sub provided) ──
        if action == "list_files":
            # Filter prefix: if user_sub is provided list only their folder
            prefix = f"uploads/user_id={user_sub}/" if user_sub else "uploads/"
            files  = []
            paginator = s3_client.get_paginator("list_objects_v2")
            for page in paginator.paginate(Bucket=BUCKET_NAME, Prefix=prefix):
                for obj in page.get("Contents", []):
                    files.append({
                        "key":           obj["Key"],
                        "name":          obj["Key"].split("/")[-1] or obj["Key"],
                        "size":          obj["Size"],
                        "last_modified": obj["LastModified"].strftime("%Y-%m-%d %H:%M UTC"),
                    })
            files.sort(key=lambda f: f["last_modified"], reverse=True)
            return resp(200, {"files": files})

        if not question:
            return resp(400, {"error": "question is required"})

        # ── Multi-file combined mode ───────────────────────────────────────────
        if s3_keys:
            per_file_limit = max(60000, MAX_FILE_CHARS // len(s3_keys))
            sections = []
            for key in s3_keys:
                try:
                    obj        = s3_client.get_object(Bucket=BUCKET_NAME, Key=key)
                    file_bytes = obj["Body"].read()
                    text       = extract_text(file_bytes, key)
                    if len(text) > per_file_limit:
                        text = text[:per_file_limit] + "\n[... truncated ...]"
                    sections.append(f"{'='*60}\nFILE: {key}\n{'='*60}\n{text}")
                except Exception as e:
                    print(f"Could not read {key}: {e}")
                    sections.append(f"{'='*60}\nFILE: {key}\n{'='*60}\n[Could not read this file]")

            answer = call_bedrock("\n\n".join(sections), question, chat_history)
            return resp(200, {"answer": answer})

        # ── Single-file mode ──────────────────────────────────────────────────
        if not s3_key:
            return resp(400, {"error": "s3_key or s3_keys is required"})

        obj        = s3_client.get_object(Bucket=BUCKET_NAME, Key=s3_key)
        file_bytes = obj["Body"].read()
        file_text  = extract_text(file_bytes, s3_key)
        answer     = call_bedrock(file_text, question, chat_history)
        return resp(200, {"answer": answer})

    except ClientError as e:
        print("Bedrock ClientError:", e)
        return resp(502, {"error": "LLM call failed", "detail": str(e)})
    except Exception as e:
        print("Error:", e)
        return resp(500, {"error": "Internal server error", "detail": str(e)})


def _resp(status: int, body: dict, headers: dict = CORS_HEADERS) -> dict:
    return {
        "statusCode": status,
        "headers":    headers,
        "body":       json.dumps(body),
    }
