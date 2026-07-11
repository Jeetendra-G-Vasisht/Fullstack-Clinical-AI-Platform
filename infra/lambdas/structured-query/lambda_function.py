"""
kBuddhi AI — structured-query Lambda
-------------------------------------------------------------
Function URL only (no API Gateway route — see chat Lambda's history
with the 29s API Gateway timeout for why).

Answers questions about a single tabular file (CSV or Excel) by having
GPT-5.5 write a SQL query against the file's real schema, executing that
query with DuckDB in-process (no external query service to wait on),
then asking GPT-5.5 to phrase the small result set as a natural-language
answer. The model only ever sees the column schema, a few sample rows,
and the final (small) query result — never the whole file — so this
can't time out or need a giant context window the way the raw-text
chat endpoint does.

Request:  { "s3_key": "uploads/user_id=.../file.csv", "question": "..." }
Response: { "answer": "...", "sql": "<the generated query>" }
"""

import csv
import io
import json
import os
import re
import urllib.error
import urllib.request

import boto3
import duckdb

BUCKET_NAME    = os.environ.get("BUCKET_NAME", "")
BUCKET_REGION  = os.environ.get("BUCKET_REGION", "us-east-2")
ALLOWED_ORIGIN = os.environ.get("ALLOWED_ORIGIN", "https://kbuddhiai.com")
OPENROUTER_KEY = os.environ.get("OPENROUTER_API_KEY", "")

TABULAR_EXTENSIONS = {"csv", "txt", "xlsx", "xls"}

FORBIDDEN_SQL_KEYWORDS = (
    "attach", "copy", "pragma", "install", "load ", "create ", "insert ",
    "update ", "delete ", "drop ", "alter ", "export", "import",
)


# ── File loading ────────────────────────────────────────────────────────────

def to_csv_path(file_bytes: bytes, filename: str) -> str:
    """Normalizes CSV or Excel bytes into a temp CSV file DuckDB can read."""
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    tmp_path = "/tmp/query_data.csv"

    if ext in ("csv", "txt"):
        # Tolerant decode — some source files have stray non-UTF-8 bytes,
        # which DuckDB's CSV reader rejects outright by default.
        text = file_bytes.decode("utf-8", errors="replace")
        with open(tmp_path, "w", newline="", encoding="utf-8") as f:
            f.write(text)
        return tmp_path

    if ext in ("xlsx", "xls"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb[wb.sheetnames[0]]
        with open(tmp_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(["" if c is None else c for c in row])
        return tmp_path

    raise ValueError(f"Unsupported file type for structured queries: .{ext}")


# ── LLM calls ───────────────────────────────────────────────────────────────

def call_llm(system_msg: str, user_msg: str, max_tokens: int) -> str:
    payload = json.dumps({
        "model": "openai/gpt-5.5",
        "messages": [
            {"role": "system", "content": system_msg},
            {"role": "user", "content": user_msg},
        ],
        "max_tokens": max_tokens,
    }).encode("utf-8")

    req = urllib.request.Request(
        "https://openrouter.ai/api/v1/chat/completions",
        data=payload,
        headers={
            "Authorization": f"Bearer {OPENROUTER_KEY}",
            "Content-Type":  "application/json",
            "HTTP-Referer":  "https://kbuddhiai.com",
            "X-Title":       "kBuddhi AI Structured Query",
        },
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=60) as resp:
        result = json.loads(resp.read().decode("utf-8"))
    return result["choices"][0]["message"]["content"]


def extract_sql(raw: str) -> str:
    if not raw:
        return ""  # LLM returned no content — caller treats this as an unsafe/empty query
    text = raw.strip()
    fence = re.search(r"```(?:sql)?\s*(.*?)```", text, re.DOTALL | re.IGNORECASE)
    if fence:
        text = fence.group(1).strip()
    return text.split(";")[0].strip()  # first statement only — no stacked queries


def is_safe_select(sql: str) -> bool:
    lowered = sql.strip().lower()
    if not (lowered.startswith("select") or lowered.startswith("with")):
        return False
    return not any(word in lowered for word in FORBIDDEN_SQL_KEYWORDS)


def generate_sql(schema_desc: str, sample_desc: str, question: str) -> str:
    system_msg = (
        "You write a single DuckDB SQL SELECT query against a table named `data`. "
        "Output ONLY the raw SQL — no explanation, no markdown fences. "
        "Double-quote any column name that contains spaces or special characters.\n\n"
        f"Table schema (column, type):\n{schema_desc}\n\nSample rows:\n{sample_desc}"
    )
    return extract_sql(call_llm(system_msg, question, max_tokens=500))


def phrase_answer(question: str, columns: list, rows: list) -> str:
    result_text = "Columns: " + ", ".join(columns) + "\nRows:\n" + "\n".join(
        str(r) for r in rows[:50]
    )
    system_msg = (
        "Answer the user's question using ONLY the query result data below. "
        "Be concise and direct.\n\n" + result_text
    )
    return call_llm(system_msg, question, max_tokens=1000)


# ── Lambda handler ─────────────────────────────────────────────────────────

def lambda_handler(event, context):
    is_function_url = event.get("version") == "2.0"
    cors = {} if is_function_url else {
        "Access-Control-Allow-Origin":  ALLOWED_ORIGIN,
        "Access-Control-Allow-Headers": "Content-Type",
        "Access-Control-Allow-Methods": "POST, OPTIONS",
    }

    def resp(status, body):
        return {"statusCode": status, "headers": cors, "body": json.dumps(body)}

    try:
        body     = json.loads(event.get("body") or "{}")
        s3_key   = body.get("s3_key", "").strip()
        question = body.get("question", "").strip()

        if not s3_key or not question:
            return resp(400, {"error": "s3_key and question are required"})

        ext = s3_key.lower().rsplit(".", 1)[-1] if "." in s3_key else ""
        if ext not in TABULAR_EXTENSIONS:
            return resp(400, {"error": f"Structured queries only support CSV/Excel files, got .{ext}"})

        s3  = boto3.client("s3", region_name=BUCKET_REGION)
        obj = s3.get_object(Bucket=BUCKET_NAME, Key=s3_key)
        file_bytes = obj["Body"].read()

        csv_path = to_csv_path(file_bytes, s3_key)

        con = duckdb.connect()
        con.execute(f"CREATE TABLE data AS SELECT * FROM read_csv_auto('{csv_path}')")
        schema_desc = "\n".join(f"{r[0]} ({r[1]})" for r in con.execute("DESCRIBE data").fetchall())
        sample_desc = "\n".join(str(r) for r in con.execute("SELECT * FROM data LIMIT 5").fetchall())

        sql = generate_sql(schema_desc, sample_desc, question)
        if not is_safe_select(sql):
            return resp(400, {"error": "Generated query was rejected for safety", "sql": sql})

        try:
            result  = con.execute(sql)
            columns = [d[0] for d in result.description]
            rows    = result.fetchall()
        except Exception as e:
            # One retry — feed the error back and ask GPT to fix the query.
            fix_prompt = f"This query failed with error: {e}\nQuery: {sql}\nFix it."
            sql2 = extract_sql(call_llm(
                f"Table schema (column, type):\n{schema_desc}\n\nSample rows:\n{sample_desc}",
                fix_prompt, max_tokens=500,
            ))
            if not is_safe_select(sql2):
                return resp(500, {"error": f"Query failed: {e}", "sql": sql})
            result  = con.execute(sql2)
            columns = [d[0] for d in result.description]
            rows    = result.fetchall()
            sql = sql2

        answer = phrase_answer(question, columns, rows)
        return resp(200, {"answer": answer, "sql": sql})

    except urllib.error.HTTPError as e:
        detail = e.read().decode("utf-8", errors="replace")
        print("OpenRouter HTTPError:", e.code, detail)
        return resp(502, {"error": "LLM call failed", "detail": detail})
    except Exception as e:
        print("Error:", e)
        return resp(500, {"error": "Internal server error", "detail": str(e)})
